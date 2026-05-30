"""Importa clientes do arquivo Marketup para a base AriOne.

Uso:
    python scripts/importar_clientes_marketup.py "C:\\Users\\Arisvan\\Downloads\\ListaClientes.xlsx"

O script suporta arquivos XLSX ou CSV e atualiza clientes existentes pelo CPF/CNPJ.
"""

import csv
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

from app import create_app, db
from app.models.cadastros.cliente import Cliente

try:
    import openpyxl
except ImportError:
    openpyxl = None


HEADER_MAP = {
    'codigosistema': None,
    'nomerazaosocial': 'nome',
    'apelidonomefantasia': 'apelido',
    'tipolistadeprecos': 'tabela_preco',
    'sexomouf': 'genero',
    'cpf': 'cpf',
    'rg': 'rg_ie',
    'expedicaoorg': None,
    'ufdorg': None,
    'indicadoriedestinatario': None,
    'cnpj': 'cnpj',
    'ie': 'rg_ie',
    'telefone': 'telefone',
    'celular': 'celular',
    'fax': None,
    'email': 'email',
    'site': None,
    'endereco': 'end_res_logradouro',
    'numero': 'end_res_numero',
    'complemento': 'end_res_complemento',
    'bairro': 'end_res_bairro',
    'cidade': 'end_res_cidade',
    'estado': 'end_res_uf',
    'cep': 'end_res_cep',
    'datadenascimento': 'data_nascimento',
}


def normalize_header(header):
    if not header:
        return ''
    header = str(header).strip().lower()
    # Remove accents
    header = ''.join(ch for ch in unicodedata.normalize('NFD', header) if unicodedata.category(ch) != 'Mn')
    # Remove non-alphanumeric
    header = ''.join(ch for ch in header if ch.isalnum())
    return header


def normalize_document(value):
    if not value:
        return ''
    digits = ''.join(ch for ch in str(value) if ch.isdigit())
    return digits


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def row_to_dict(row, headers):
    mapped = {}
    for key, value in zip(headers, row):
        if not key:
            continue
        field = HEADER_MAP.get(key)
        if field is None:
            continue
        mapped[field] = value
    return mapped


def read_xlsx(path):
    if openpyxl is None:
        raise RuntimeError('openpyxl não está instalado. Rode pip install openpyxl')

    wb = openpyxl.load_workbook(path, data_only=True)
    if not wb.sheetnames:
        raise RuntimeError('Arquivo XLSX sem abas válidas.')

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(cell) for cell in rows[0]]
    return [row_to_dict(row, headers) for row in rows[1:]]


def read_csv(path):
    with path.open('r', encoding='utf-8-sig', newline='') as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        headers = [normalize_header(h) for h in reader.fieldnames or []]
        result = []
        for row in reader:
            normalized = {normalize_header(k): v for k, v in row.items()}
            mapped = {HEADER_MAP.get(k): v for k, v in normalized.items() if HEADER_MAP.get(k) is not None}
            result.append(mapped)
        return result


def import_clients(path):
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f'Arquivo não encontrado: {path}')

    if path.suffix.lower() == '.xlsx':
        rows = read_xlsx(path)
    elif path.suffix.lower() in ('.csv', '.txt'):
        rows = read_csv(path)
    else:
        raise RuntimeError('Formato não suportado. Use XLSX ou CSV.')

    total_new = 0
    total_updated = 0
    total_skipped = 0

    for row in rows:
        cpf = normalize_document(row.get('cpf'))
        cnpj = normalize_document(row.get('cnpj'))
        cpf_cnpj = cnpj or cpf
        nome = str(row.get('nome') or '').strip()
        
        if not cpf_cnpj or not nome:
            total_skipped += 1
            continue

        tipo_pessoa = 'J' if cnpj else 'F'
        cliente = Cliente.query.filter_by(cpf_cnpj=cpf_cnpj).first()
        if cliente is None:
            cliente = Cliente(cpf_cnpj=cpf_cnpj, ativo=True)
            total_new += 1
        else:
            total_updated += 1

        cliente.tipo_pessoa = tipo_pessoa
        cliente.nome = nome
        cliente.apelido = str(row.get('apelido') or '').strip() or cliente.apelido
        cliente.rg_ie = str(row.get('rg_ie') or row.get('ie') or '').strip() or cliente.rg_ie
        cliente.genero = str(row.get('genero') or '').strip() or cliente.genero
        cliente.email = str(row.get('email') or '').strip() or cliente.email
        cliente.tabela_preco = str(row.get('tabela_preco') or '').strip() or cliente.tabela_preco
        cliente.end_res_logradouro = str(row.get('end_res_logradouro') or '').strip() or cliente.end_res_logradouro
        cliente.end_res_numero = str(row.get('end_res_numero') or '').strip() or cliente.end_res_numero
        cliente.end_res_complemento = str(row.get('end_res_complemento') or '').strip() or cliente.end_res_complemento
        cliente.end_res_bairro = str(row.get('end_res_bairro') or '').strip() or cliente.end_res_bairro
        cliente.end_res_cidade = str(row.get('end_res_cidade') or '').strip() or cliente.end_res_cidade
        cliente.end_res_uf = str(row.get('end_res_uf') or '').strip() or cliente.end_res_uf
        cliente.end_res_cep = str(row.get('end_res_cep') or '').strip() or cliente.end_res_cep
        cliente.data_nascimento = parse_date(row.get('data_nascimento')) or cliente.data_nascimento

        celular = str(row.get('celular') or '').strip()
        telefone = str(row.get('telefone') or '').strip()
        if celular:
            cliente.whatsapp = celular
        if telefone:
            cliente.telefone = telefone
        if not cliente.whatsapp and telefone:
            cliente.whatsapp = telefone

        db.session.add(cliente)

    db.session.commit()
    return total_new, total_updated, total_skipped


def main():
    if len(sys.argv) != 2:
        print('Uso: python scripts/importar_clientes_marketup.py "C:\\Users\\Arisvan\\Downloads\\ListaClientes.xlsx"')
        sys.exit(1)

    path = sys.argv[1]
    app = create_app()
    with app.app_context():
        new_count, updated_count, skipped_count = import_clients(path)
        print(f'Importação finalizada. Novos: {new_count}, Atualizados: {updated_count}, Pulados (sem CPF/CNPJ): {skipped_count}')


if __name__ == '__main__':
    main()
