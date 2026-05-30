"""
Script para extrair Email, Data de Nascimento/Abertura de ListaClientes.xlsx
e adicionar ao banco de dados de Clientes
"""

import sys
from pathlib import Path
from datetime import datetime

# Adiciona o caminho do projeto
sys.path.insert(0, str(Path(__file__).parent.parent))

from app import create_app, db
from app.models.cadastros.cliente import Cliente
import openpyxl

def parse_date(value):
    """Converte data em vários formatos"""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    
    text = str(value).strip()
    if not text:
        return None
    
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d.%m.%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None

def normalize_doc(value):
    """Remove caracteres não numéricos"""
    if not value:
        return ''
    return ''.join(c for c in str(value) if c.isdigit())

def main():
    app = create_app()
    with app.app_context():
        print("📊 Lendo ListaClientes.xlsx...")
        
        wb = openpyxl.load_workbook(r'C:\Users\Arisvan\Downloads\ListaClientes.xlsx', data_only=True)
        ws = wb[wb.sheetnames[0]]
        
        # Índices de colunas (0-based)
        COLS = {
            'nome': 1,              # B: Nome/Razão Social
            'tipo': 3,              # D: Tipo (Lista de Preços)
            'cpf': 5,               # F: CPF
            'cnpj': 10,             # K: CNPJ
            'tipo_pessoa': 4,       # E: Sexo/Tipo
            'email': 15,            # P: Email
            'site': 16,             # Q: Site
            'data_nasc': 24,        # Y: Data de Nascimento
            'telefone': 12,         # M: Telefone
            'celular': 13,          # N: Celular
        }
        
        total = 0
        atualizado = 0
        adicionado = 0
        erros = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                nome = (row[COLS['nome']] or '').strip()
                if not nome:
                    continue
                
                # Determina CPF/CNPJ
                cpf_str = normalize_doc(row[COLS['cpf']])
                cnpj_str = normalize_doc(row[COLS['cnpj']])
                
                if not cpf_str and not cnpj_str:
                    print(f"  ⚠️  Linha {row_num}: Sem CPF/CNPJ, pulando...")
                    continue
                
                # Usa CNPJ se existir, senão CPF
                doc = cnpj_str if cnpj_str else cpf_str
                tipo_pessoa = 'J' if cnpj_str else 'F'
                
                # Busca cliente existente
                cliente = Cliente.query.filter_by(cpf_cnpj=doc).first()
                
                if cliente:
                    # Atualiza dados
                    cliente.email = (row[COLS['email']] or '').strip() or cliente.email
                    cliente.site = (row[COLS['site']] or '').strip() or getattr(cliente, 'site', None)
                    
                    if tipo_pessoa == 'F':
                        data_n = parse_date(row[COLS['data_nasc']])
                        if data_n:
                            cliente.data_nascimento = data_n
                    else:
                        # Para PJ, seria data_abertura (campo que ainda precisa ser criado)
                        data_a = parse_date(row[COLS['data_nasc']])
                        if data_a and hasattr(cliente, 'data_abertura'):
                            cliente.data_abertura = data_a
                    
                    cliente.telefone = (row[COLS['telefone']] or '').strip() or cliente.telefone
                    cliente.whatsapp = (row[COLS['celular']] or '').strip() or cliente.whatsapp
                    
                    db.session.add(cliente)
                    atualizado += 1
                    print(f"  ✏️  Linha {row_num}: Atualizado {nome}")
                else:
                    # Cria novo cliente
                    cliente = Cliente(
                        nome=nome,
                        cpf_cnpj=doc,
                        tipo_pessoa=tipo_pessoa,
                        email=(row[COLS['email']] or '').strip() or None,
                        telefone=(row[COLS['telefone']] or '').strip() or None,
                        whatsapp=(row[COLS['celular']] or '').strip() or None,
                        ativo=True,
                    )
                    
                    # Tenta adicionar site se disponível
                    if hasattr(cliente, 'site'):
                        cliente.site = (row[COLS['site']] or '').strip() or None
                    
                    # Adiciona data conforme tipo
                    if tipo_pessoa == 'F':
                        data_n = parse_date(row[COLS['data_nasc']])
                        if data_n:
                            cliente.data_nascimento = data_n
                    else:
                        # Para PJ, seria data_abertura
                        if hasattr(cliente, 'data_abertura'):
                            data_a = parse_date(row[COLS['data_nasc']])
                            if data_a:
                                cliente.data_abertura = data_a
                    
                    db.session.add(cliente)
                    adicionado += 1
                    print(f"  ✅ Linha {row_num}: Adicionado {nome}")
                
                total += 1
                
            except Exception as e:
                erros += 1
                print(f"  ❌ Linha {row_num}: Erro - {str(e)}")
        
        # Commit
        try:
            db.session.commit()
            print(f"\n✅ IMPORTAÇÃO CONCLUÍDA:")
            print(f"   Total processados: {total}")
            print(f"   Adicionados: {adicionado}")
            print(f"   Atualizados: {atualizado}")
            print(f"   Erros: {erros}")
        except Exception as e:
            db.session.rollback()
            print(f"\n❌ ERRO ao salvar: {str(e)}")

if __name__ == '__main__':
    main()
